
# ──────────────────────────────────────────────────────────────
# 📦 DICOCH DICOM Converter – 개발 이력 / Changelog
# 개발자: 송정일 (Song Jung-il), 국립문화재연구원 X-ray/CT 분석실
# GitHub : https://github.com/SONG-JUNG-IL/DICOCH_TAG
# 배포 라이선스: CC BY-SA 4.0
# ──────────────────────────────────────────────────────────────
# ✅ v3.2 · 2025-07-25
#   - dicom.dic → Dic_DICOCH_DICOM_PravateTag_YYYYMMDD_HHMMSS.txt 자동 저장
#   - 태그 유효성 검사 결과를 error 컬럼으로 GUI 및 TXT/XLSX/JSON에 병합 저장
#   - IIIF manifest URL 자동 탐색 및 Mirador 연결 기능 강화
#   - 변환 결과 로그 log_YYYYMMDD_HHMMSS.txt 자동 저장
#   - 변환 완료 후 출력 폴더 자동 열기 + IIIF 뷰어 실행 옵션 추가
#   - GUI 출력 로그 개선 및 _update_tag_view 리팩토링
#
# 📌 이전 버전 요약
# - v3.1 (2025-06-24): RescaleSlope/Intercept GUI 적용, JSON 저장 지원
# - v3.0 (2025-06-23): Info 탭 UI 개선, Creator 중복 해결
# - v2.9u (2025-06-22): IIIF 뷰어 옵션 추가, dicom.dic 재작업
# - v2.9s (2025-06-20): Excel 기반 Slope/Intercept 우선, 로그 및 저장 기능 추가
# ──────────────────────────────────────────────────────────────

#* **v3.1 · 2025-06-24**

#  * RescaleSlope/Intercept “GUI 우선 적용” 체크박스 완성

#    * `build_dataset()`에 `override` 파라미터 추가
#    * `_convert()`에서 `self.gui_override.get()` 전달
#  * 태그 결과 **JSON** 저장 지원 (TXT / XLSX / JSON 3-way)
#  * `_update_tag_view()` 위젯 재생성 코드 삭제로 `NameError` 해결

#* **v3.0 · 2025-06-23**

#  * Info 탭 UI 개선 (Segoe UI 11pt, 헤더 Bold, 행간 2px)
#  * Creator 중복 해결: `insert_block_creator()` 로 블록별 1줄만 삽입
#  * CREATORS 토큰(0x10\~0x18) 재정렬 (≤16 byte)

#* **v2.9u · 2025-06-22**

#  * IIIF auto-viewer 옵션 추가
#  * `dicom.dic` 재작업 (Private Tag 매핑)

#* **v2.9s · 2025-06-20**

#  * RescaleSlope/Intercept: 엑셀 값 우선, 없으면 GUI 값 보완
#  * 변환 로그 뷰 + 태그 결과 저장 (TXT·XLSX) 기능
#  * UR→UT 변환, 빈 SQ prune, UTF-8, `ThreadPoolExecutor` 유지

#---

## English Version

# ──────────────────────────────────────────────────────────────
# 📦 DICOCH DICOM Converter – Changelog
# Author: Song Jung-il (송정일), X-ray/CT Lab, NRICH (National Research Institute of Cultural Heritage)
# GitHub: https://github.com/SONG-JUNG-IL/DICOCH_TAG
# License: CC BY-SA 4.0 — Free use, modification and redistribution with attribution
# ──────────────────────────────────────────────────────────────
# ✅ v3.2 · 2025-07-25
#   - Auto-renamed dicom.dic to Dic_DICOCH_DICOM_PravateTag_YYYYMMDD_HHMMSS.txt
#   - Merged validation errors into GUI display and tag exports (TXT / XLSX / JSON)
#   - Improved auto-detection of IIIF manifest URLs and Mirador viewer integration
#   - Automatically saved conversion log as log_YYYYMMDD_HHMMSS.txt
#   - Added options to auto-open output folder and launch IIIF viewer after conversion
#   - Refactored _update_tag_view and enhanced log output formatting
#
# 📌 Previous Versions
# - v3.1 (2025-06-24): Applied GUI-first RescaleSlope/Intercept, added JSON export
# - v3.0 (2025-06-23): Enhanced Info tab UI, fixed Creator duplication logic
# - v2.9u (2025-06-22): Added IIIF viewer option, restructured dicom.dic mapping
# - v2.9s (2025-06-20): Prioritized Excel-based slope/intercept, enabled tag export and logging
# ──────────────────────────────────────────────────────────────

### Changelog

#* **v3.1 · 2025-06-24**

#  * Completed “GUI-first” RescaleSlope/Intercept checkbox

#    * Added `override` parameter to `build_dataset()`
#    * Passed `self.gui_override.get()` into `_convert()`
#  * Added **JSON** export for tag results (TXT / XLSX / JSON 3-way)
#  * Removed widget re-creation in `_update_tag_view()` to fix `NameError`

#* **v3.0 · 2025-06-23**

#  * Improved Info-tab UI (Segoe UI 11pt, bold headers, 2px line spacing)
#  * Resolved Creator duplication via `insert_block_creator()` (one-line per block)
#  * Reordered CREATORS tokens (0x10–0x18) to ≤16 bytes

#* **v2.9u · 2025-06-22**

#  * Added IIIF auto-viewer option
#  * Reworked `dicom.dic` for Private Tag mapping

#* **v2.9s · 2025-06-20**

#  * RescaleSlope/Intercept: prefer Excel values, fallback to GUI inputs
# * Conversion log viewer + tag result export (TXT·XLSX)
#  * UR→UT mapping, empty SQ pruning, UTF-8 support, `ThreadPoolExecutor`

# ──────────────────────────────────────────────────────────────

from __future__ import annotations
import json, os, re, threading, concurrent.futures as cf, sys, subprocess, webbrowser
from datetime import datetime as dt
from pathlib import Path
from typing import Callable, Dict, List, Set
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import numpy as np
import pydicom
from pydicom.dataset import Dataset, FileDataset
from pydicom.sequence import Sequence
from pydicom.uid import generate_uid, SecondaryCaptureImageStorage, ExplicitVRLittleEndian
from pydicom.datadict import add_private_dict_entry
from PIL import Image
import tifffile as tiff
from openpyxl import load_workbook
import logging, re
import pandas as pd


# ── VR 규격 정보 ────────────────────────────────────────────
VR_MAXLEN = {
    "AE":16,"AS":4,"CS":16,"DA":8,"DS":16,"DT":26,"IS":12,
    "LO":64,"LT":10240,"PN":64,"SH":16,"ST":1024,"TM":16,
    "UI":64,"UT":0xFFFFFFFE,
   # ── Numeric VRs (추가) ───────────────────────
    "UL":12,  # Unsigned Long
    "US":12,  # Unsigned Short
    "SL":12,  # Signed  Long
    "SS":12,  # Signed  Short
    "FL":16,  # Float 32-bit
    "FD":32   # Float 64-bit
    }
# ───────────────────────────────────────────
# ───────────────────────────────────────────────────────────
#  기본 VR 사전  (key = (Group, Element)  , value = VR String)
#  ※ 한 번만 정의, 전역에서 재사용
# ───────────────────────────────────────────────────────────
VR_DEF: dict[tuple[str, str], str] = {
    # ── 식별 / 날짜·시간 / UID ──────────────────────────
    ("0008","0060"): "CS",   # Modality
    ("0008","0020"): "DA",   # StudyDate
    ("0008","0030"): "TM",   # StudyTime
    ("0008","0018"): "UI",   # SOPInstanceUID
    ("0020","000D"): "UI",   # StudyInstanceUID
    ("0020","000E"): "UI",   # SeriesInstanceUID
    ("0020","0052"): "UI",   # FrameOfReferenceUID

    # ── 이미지 픽셀 메타 ───────────────────────────────
    ("0028","0002"): "US",   # SamplesPerPixel
    ("0028","0004"): "CS",   # PhotometricInterpretation
    ("0028","0010"): "US",   # Rows
    ("0028","0011"): "US",   # Columns
    ("0028","0100"): "US",   # BitsAllocated
    ("0028","0101"): "US",   # BitsStored
    ("0028","0102"): "US",   # HighBit
    ("0028","0103"): "US",   # PixelRepresentation
    # ("0028","0006"): "US", # PlanarConfiguration (필요 시 활성)

    # ── Secondary Capture 권장 셋 ──────────────────────
    ("0008","0008"): "CS",   # ImageType  (DERIVED\SECONDARY)
    ("0008","0070"): "LO",   # Manufacturer
    ("0008","1090"): "LO",   # ManufacturerModelName
}

# ───────────────────────────────────────────

VALID_VR = set(VR_MAXLEN) | {"SQ","UN","OW","OB","OF","OD","OL","UR"}

# ── 정규식 · 헬퍼 ───────────────────────────────────────────
HEX_RE      = re.compile(r"[0-9A-Fa-f]{1,4}$")
NUM_RE = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")   # ① 숫자 패턴
MULTI_DELIM = re.compile(r"[,;/\s]+")    
# ── 상수 ───────────────────────────────────────────────
MAX_UT   = 0xFFFFFFFE
GROUP_HEX = "0013"
VR_DOWN  = {"UR": "UT"}          # UR→UT

# ──로깅 기본 설정───────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
LOG = logging.getLogger("dicoch.loader")

# ── ▣ 기관용 OID 프리픽스 설정 예시  (Korean)──────────────────────────
#     ※ 실제 배포 전에는 다음 두 가지 경로 중 하나를 통해
#        ‘공식’ 기관 OID 루트를 발급받아 아래 값을 교체해야 합니다.
#
#        1) **NEMA 등록 (권장)**  
#           · DICOM 표준 주관 단체(NEMA) 무료 신청  
#           · `1.2.840.xxxxxxx…` 형식의 고유 Prefix 부여
#
#        2) 대한민국 OID 관리기관(KATS) 신청  
#           · `1.2.410.기관번호…` 형식
#
#   예시 루트 : 1.2.410.999999.20250724.
#   └──┬──┬──┬───────────┘
#      │  │  └─ 999999 : (임시) 기관 고유 번호
#      │  └──── 410     : “대한민국” 국가 OID
#      └──────── 1.2    : ISO(International) OID 트리
#
#   ▸ 20250724 부분은 발행 날짜·버전 구분 용도로 붙인 임의 숫자입니다.
#   ▸ UID 전체는 “숫자와 점( . )”만 사용, 64 byte 이하 조건을 반드시 지켜야 합니다.
#
#  ▣ Example of Institutional OID Prefix  (English)
#     ※ Before production release, replace this with an **official**
#       root OID obtained via one of the following:
#
#        1) **NEMA registration (recommended)**  
#           · Free application to the DICOM Standards Committee (NEMA)  
#           · Official prefix in the form `1.2.840.xxxxxxx…`
#
#        2) Korean OID authority (KATS)  
#           · Prefix in the form `1.2.410.<organization-id>…`
#
#   Sample root : 1.2.410.999999.20250724.
#   └──┬──┬──┬───────────┘
#      │  │  └─ 999999 : (temporary) organization-specific number
#      │  └──── 410     : Country OID for **Republic of Korea**
#      └──────── 1.2    : ISO/ITU-T OID tree
#
#   ▸ The trailing “20250724” is just an example timestamp / version tag.
#   ▸ A valid DICOM UID must contain only digits and dots, and be ≤ 64 bytes long.
# ──────────────────────────────────────────────────────────────
OID_PREFIX = "1.2.410.999999.20250724."




### ▶ CHG: CREATORS 9 종 — 엑셀과 동일한 이름으로 교체
CREATORS = [
    "DICOCH",                # 0x0010  (루트 Creator)
    "Heritage_NRICH",        # 0x0011  Heritage Metadata Seq.
    "GrayCal_NRICH",         # 0x0012  Gray-Value Calibration Seq.
    "ROIGray_NRICH",         # 0x0013  ROI Gray-Value Seq.
    "HUCal_NRICH",           # 0x0014  HU Calibration Seq.
    "ROIHU_NRICH",           # 0x0015  ROI-HU Seq.
    "IIIF_NRICH",            # 0x0016  IIIF / Link Seq.
    "Security_NRICH",        # 0x0017  Security / Access Seq.
    "DICOCH_Dict_NRICH",     # 0x0018  Private-Tag Dictionary Seq.
]

MIRADOR_DEMO = "https://projectmirador.org/demo/?manifest="

# ── VR 보정 ────────────────────────────────────────────
def _num(s, default="0"):
    """임의 문자열 s → 첫 실수/정수 문자열; 없으면 default"""
    m = NUM_RE.search(str(s))
    return m.group(0) if m else default                  # ① 수정
def _fix_cs(v): return str(v).upper().replace(" ", "_")[:16]
def _fix_da(v): return re.sub(r"\D", "", str(v))[:8]
def _fix_tm(v): return (re.sub(r"\D", "", str(v)) + "000000")[:6]

# VR_RULES – 중복 PN 제거 및 정렬
VR_RULES: Dict[str, Callable[[str], object]] = {
    "CS": _fix_cs,
    "PN": lambda v: str(v).replace(",", "^")[:64],        # ③ 하나만 유지
    "SH": lambda v: str(v)[:16],
    "LO": lambda v: str(v)[:64],
    "UI": lambda v: str(v)[:64],
    "UT": lambda v: v,
    "DS": lambda v: _num(v)[:16],
    "IS": lambda v: str(int(float(_num(v)))),
    "US": lambda v: int(float(_num(v))),
    "UL": lambda v: int(float(_num(v))),
    "FL": lambda v: float(_num(v)),
    "FD": lambda v: float(_num(v)),
    "OB": lambda v: v,
    "OW": lambda v: v,
    "DA": _fix_da,
    "TM": _fix_tm,
}

# --------------------------------------------------------------------------
# ── DataFrame 헤더 표준화 공통 함수 ─────────────────────────
def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    """소문자+공백 제거 + 중복 열 첫 번째만 유지 (별칭은 만들지 않음)"""
    df.columns = [c.lower().strip().replace(" ", "") for c in df.columns]
    return df.loc[:, ~df.columns.duplicated(keep="first")]


# OB / OW / UN VR 안전 변환
def safe_value(vr: str, val: str):
    if vr in ("OB", "OW", "UN"):
        v = str(val).strip()
        try:
            return bytes.fromhex(v)               # hex 입력 지원
        except ValueError:
            return v.encode("utf-8")              # 일반 문자열

    if vr == "UI":                     # UID 정규화
        uid = _clean_ui(val)
        return uid

    return VR_RULES.get(vr, lambda x: x)(val)

# ── Creator 삽입 & 사전 등록 ───────────────────────────
# ---------- Creator helpers (루트·SQ 전용) ----------
def insert_all_creators(ds: Dataset):
    """루트 Dataset에 8개 Creator 전부 삽입"""
    for slot, name in enumerate(CREATORS):
        ds.add_new((0x0013, 0x0010 + slot), "LO", name)

def insert_block_creator(ds: Dataset, block_hex: str):
    """SQ-Item 안에 DICOCH + 해당 블록 Creator 한 줄만 삽입"""
    ds.add_new((0x0013, 0x0010), "LO", "DICOCH")      # 기본 Creator
    bb = int(block_hex, 16) - 0x10                    # 0x11→1 … 0x17→7
    if 0 <= bb < len(CREATORS):
        ds.add_new((0x0013, 0x0010 + bb), "LO", CREATORS[bb])

def register_private_tags(df: pd.DataFrame):
    for slot, name in enumerate(CREATORS):
        add_private_dict_entry(name, 0x00130010 + slot, "LO", name, "1")
    for _, r in df.iterrows():
        if r["element"] == "0010": 
            continue
        tag_int = (int(GROUP_HEX,16)<<16) | int(r["element"],16)
        slot = max(0, min(7, int(r["element"][:2],16)-0x10))
        add_private_dict_entry(
            CREATORS[slot], tag_int,
            VR_DOWN.get(r["vr"], r["vr"]),
            r["keyword"] or f"DICOCH_{r['element']}", "1"
        )

# ── dicom.dic ─────────────────────────────────────────
def write_dic(df: pd.DataFrame, out_dir: Path) -> Path:
    #df = _normalize_cols(df)          # ← 추가
    # 1 ) 소문자 표준화
    df.columns = [c.lower().strip().replace(" ", "") for c in df.columns]
    
    # 2 ) ★ 중복된 열은 첫 번째 것만 남기기
    df = df.loc[:, ~df.columns.duplicated(keep="first")]

    out_dic = out_dir / "dicom.dic"
    lines   = []

    for _, r in df.iterrows():
        if r["vr"] == "SQ":
            continue
        vr = VR_DOWN.get(r["vr"], r["vr"])
        kw = r["keyword"] or f"DICOCH_{r['element']}"
        lines.append(f"({r['group']},{r['element']}) {vr} 1 {kw}")



    out_dic.write_text("\n".join(lines), encoding="utf-8")
    return out_dic




# ── 태그 로더 ──────────────────────────────────────────
def _parse_elem(e):
    if pd.isna(e) or str(e).strip() == "":
        return ""
    s=str(e).strip()
    return s.zfill(4).upper() if re.fullmatch(r"[0-9A-Fa-f]{1,4}",s) else f"{int(float(s)):04X}"

# ── 공통 헬퍼 ───────────────────────────────────────────
def _parse_hex(v:str)->str:
    s = str(v).strip()
    if not s: return ""
    if HEX_RE.fullmatch(s): return s.zfill(4).upper()
    return f"{int(float(s)):04X}"

def _truncate(v:str, vr:str)->str:
    lim = VR_MAXLEN.get(vr, 1024)
    return v[:lim] if len(v) > lim else v

# UI 정규화 (점 없는 숫자 → 루트 접두사)
def _clean_ui(v: str) -> str:
    """UID 문자열 정규화
       1) 숫자·점만 남김
       2) 선행 0 제거
       3) 점이 하나도 없으면 OID_PREFIX 붙이기
       4) 끝에 점 있으면 삭제
       5) 64 byte 초과 시 truncate
    """
    raw = re.sub(r"[^\d.]", "", str(v))          # 1)
    if "." not in raw:                           # 3)
        raw = f"{OID_PREFIX}{raw}"
    if raw.endswith("."):                        # 4)
        raw = raw[:-1]
    # 2) 각 arc 의 선행 0 제거 (단, "0"은 유지)
    arcs = [str(int(a)) if a and a != "0" else "0" for a in raw.split(".")]
    uid  = ".".join(arcs)
    return _truncate(uid, "UI")                   # 5)

def _clean_ds(v):
    """DS 다중 값 구분자를 ‘\’(백슬래시)로 표준화"""
    return MULTI_DELIM.sub(r"\\", str(v).strip("[] "))  

def _clean_cs(v:str)->str:
    return re.sub(r"[^A-Z0-9_]", "", v.upper())

def _add_if_missing(df: pd.DataFrame,
                    g: str, e: str, vr: str, value: str) -> None:
    """시트에 (g,e)가 없으면 새 행 추가, 있으면 빈 value 보완."""
    sel = (df["group"] == g) & (df["element"] == e)
    if sel.any():
        df.loc[sel & df["value"].eq(""), "value"] = value
    else:                                    # 새 행 추가
        df.loc[len(df)] = {
            "group": g, "element": e, "vr": vr,
            "keyword": "", "value": value, "parenttag": ""
        }

def _hex2int(s: str) -> int:
    """빈 문자열이면 -1, 아니면 16진수 → int"""
    return int(s, 16) if s else -1

# ── 최종 load_tags() ────────────────────────────────────────
def load_tags(xlsx_path:str|Path)->pd.DataFrame:
    # ① 엑셀 로드 & 헤더 정규화
    df = pd.read_excel(xlsx_path, dtype=str).fillna("")
    df.columns = [c.lower().strip().replace(" ", "") for c in df.columns]
    # ★ parenttag 열이 없으면 자동으로 추가 ——▼
    if "parenttag" not in df.columns:
        df["parenttag"] = ""
    # ② 공백 제거 · HEX 통일 · VR 대문자
    for col in ("keyword","value","vr","parenttag"):
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    df["group"]   = df["group"].apply(_parse_hex)
    df["element"] = df["element"].apply(_parse_hex)
    df["vr"]      = df["vr"].str.upper().str.strip()

    # ③ VR 미인식 → LO
    bad = df[~df["vr"].isin(VALID_VR)]
    if not bad.empty:
        LOG.warning("Unknown VR → LO :\n%s", bad[["group","element","vr"]])
        df.loc[bad.index, "vr"] = "LO"

    # ④ VR-별 Value 클린업
    for i,r in df.iterrows():
        vr, val = r["vr"], r["value"]
        if vr=="UI": df.at[i,"value"] = _truncate(_clean_ui(val),"UI")
        elif vr=="DS": df.at[i,"value"] = _truncate(_clean_ds(val),"DS")
        elif vr=="IS":
            m = NUM_RE.search(val)
            if m: df.at[i,"value"] = str(int(round(float(m.group(0)))))
        elif vr=="CS": df.at[i,"value"] = _truncate(_clean_cs(val),"CS")
        elif vr=="DA": df.at[i,"value"] = _fix_da(val)
        elif vr=="TM": df.at[i,"value"] = _fix_tm(val)
        # ── 잘못된 parenttag 자동 정리 ─────────────────────────
        def _parenttag_valid(df: pd.DataFrame, tag: str) -> bool:
            sq_mask = (
                (df["group"] == tag[:4]) &
                (df["element"] == tag[4:]) &
                (df["vr"].str.upper() == "SQ")
            )
            return sq_mask.any()

        for i, r in df.iterrows():
            pt = r["parenttag"]
            # 그룹 길이(0000) 또는 SQ 태그 미존재 → 공란 처리
            if pt and (pt.endswith("0000") or not _parenttag_valid(df, pt)):
                LOG.warning("%s,%s invalid parenttag=%s → cleared",
                            r["group"], r["element"], pt)
                df.at[i, "parenttag"] = ""


    # ── ★ 공란 Group / Element 행 제거 ───────────────────
    before = len(df)
    df = df[(df["group"] != "") & (df["element"] != "")]
    if (rm := before - len(df)):
        LOG.warning("❎ 빈 Group/Element 행 %d 개 삭제", rm)

    # ⑤ 중복 태그 해소 (첫 행 유지)
    df = (df.sort_values(["group","element"])
            .groupby(["group","element","parenttag"], as_index=False)
            .first())


    # ⑥ 필수 · 권고 태그 보완
    today, now = dt.now().strftime("%Y%m%d"), dt.now().strftime("%H%M%S")
    def _add(g,e,vr,v):
        sel = (df["group"]==g) & (df["element"]==e)
        if sel.any():
            df.loc[sel & df["value"].eq(""), "value"] = v
        else:
            df.loc[len(df)] = {
                "group":g,"element":e,"vr":vr,
                "keyword":"","value":v,"parenttag":""
            }

    core = [
        ("0008","0060","CS","OT"),
        ("0010","0010","PN","UNKNOWN^HeritageObject"),
        ("0010","0020","LO","OBJ-0001"),
        ("0008","0020","DA",today),
        ("0008","0030","TM",now),
        ("0020","000D","UI",generate_uid(prefix=OID_PREFIX)),
        ("0020","000E","UI",generate_uid(prefix=OID_PREFIX)),
        ("0020","0052","UI",generate_uid(prefix=OID_PREFIX)),
    ]
    extra = [
        ("0008","0064","CS","WSD"),
        ("0008","103E","LO","Cultural Heritage Scan"),
        ("0008","0070","LO","DICOCH"),
        ("0008","1090","LO","Converter v3.2"),
        ("0008","0061","CS","OT"),
        ("0008","0018","UI",generate_uid(prefix=OID_PREFIX)),  # SOP Instance UID
    ]
    for g,e,vr,v in core+extra: _add(g,e,vr,v)

    # ⑦ 길이 초과 최종 truncate
    for i,r in df.iterrows():
        df.at[i,"value"] = _truncate(r["value"], r["vr"])

    # ⑧ 정렬 후 반환
    return df.sort_values(["group","element","parenttag"]).reset_index(drop=True)
# ──────────────────────────────────────────────────────────────

# ── TIFF 16-bit 로더 ─────────────────────────────────
def read_tiff16(p:Path)->np.ndarray:
    try:  
        return np.asarray(Image.open(p).convert("I;16"), dtype=np.uint16)
    except Exception:
        return tifffile.imread(str(p)).astype(np.uint16)

def normalize_excel_headers(xlsx_path: Path) -> None:
    """
    첫 번째 시트 1행 헤더를 모두
    → 소문자 + 선행·후행 공백 제거 + 중간 공백 제거
    ex) 'Group ' → 'group', ' Parent Tag' → 'parenttag'
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for cell in ws[1]:                                        # 1행 = 헤더
        if cell.value:
            cell.value = str(cell.value).lower().strip().replace(" ", "")
    wb.save(xlsx_path)
# ──────────────────────────────────────────────────────


# ── Dataset 빌더 ─────────────────────────────────────
# ── _has ------------------------------------------------
def _has(df, grp: str, elm: str) -> bool:
    return ((df["group"] == grp) &
            (df["element"] == elm) &
            (df["value"]   != "")).any()


# ── helper: Sequence 재귀 생성 ────────────────────────────
def build_sequence(parent_key: str, tags: pd.DataFrame) -> Sequence:
    children = tags[tags["parenttag"] == parent_key]
    items = []
    for _, r in children.iterrows():
        ds_item = Dataset()
        tag_tuple = (int(r["group"], 16), int(r["element"], 16))
        if r["vr"] == "SQ":
            sub_seq = build_sequence(r["group"] + r["element"], tags)
            if sub_seq:
                ds_item.add_new(tag_tuple, "SQ", sub_seq)
        else:
            vr = VR_DOWN.get(r["vr"], r["vr"])
            val = safe_value(r["vr"], r["value"])
            ds_item.add_new(tag_tuple, vr, val)
        items.append(ds_item)
    return Sequence(items)




# ── Rescale 파라미터 추출 ────────────────────────────────────────
def get_rescale_params(
    tags: pd.DataFrame,
    gui_slope: float,
    gui_int: float,
    override: bool
) -> tuple[str, str]:
    tags = _normalize_cols(tags)
    def _val(code: str) -> str:
        sel = (
            (tags["group"] == "0028") &
            (tags["element"] == code) &
            (tags["parenttag"] == "")
        )
        row = tags[sel]
        if row.empty:
            return ""
        v = row["value"].iloc[0]
        return str(v) if v else ""
    slope_val = _val("1053")
    int_val   = _val("1052")
    if override or not slope_val:
        slope_val = str(gui_slope)
    if override or not int_val:
        int_val = str(gui_int)
    return slope_val, int_val


# ── DICOM Dataset 생성 ─────────────────────────────────────────
# ──────────────────────────────────────────────────────────
#  DICOM Dataset 생성  (중복·들여쓰기·예외 완전 정리판)
# ──────────────────────────────────────────────────────────
def _hex2int(s: str) -> int:
    """빈 문자열이면 -1, 아니면 16진수를 int 로 변환"""
    return int(s, 16) if s else -1

def build_dataset(
    img: Path,
    tags: pd.DataFrame,
    gui_slope: float,
    gui_int: float,
    override: bool
) -> FileDataset:

    tags = _normalize_cols(tags)
    item_cache: dict[tuple[int,int], Dataset] = {}   # ★ 캐시
    seq_elem = None
    # 1) 픽셀 데이터 ---------------------------------------------------
    arr          = read_tiff16(img)
    rows, cols   = arr.shape

    # 2) File-Meta -----------------------------------------------------
    meta = Dataset()
    meta.MediaStorageSOPClassUID    = SecondaryCaptureImageStorage
    meta.MediaStorageSOPInstanceUID = generate_uid(prefix=OID_PREFIX)
    meta.TransferSyntaxUID          = ExplicitVRLittleEndian

    ds = FileDataset(img.stem + ".dcm", {}, file_meta=meta, preamble=b"\0"*128)
    ds.SOPClassUID    = meta.MediaStorageSOPClassUID
    ds.SOPInstanceUID = meta.MediaStorageSOPInstanceUID
    # ── File-Meta 필수 항목 보강
    meta.FileMetaInformationVersion = b"\x00\x01"
    meta.ImplementationClassUID     = f"{OID_PREFIX}1"
    meta.ImplementationVersionName  = "DICOCH_3_1"

    ds.SpecificCharacterSet = "ISO_IR 192"
    ds.StudyInstanceUID  = generate_uid(prefix=OID_PREFIX)
    ds.SeriesInstanceUID = generate_uid(prefix=OID_PREFIX)
    ds.ImageType         = ["DERIVED","SECONDARY"]

    now = dt.now()
    ds.StudyDate = ds.SeriesDate = ds.ContentDate = now.strftime("%Y%m%d")
    ds.StudyTime = ds.SeriesTime = ds.ContentTime = now.strftime("%H%M%S")

    ds.Modality               = "OT"
    ds.Rows, ds.Columns       = rows, cols
    ds.SamplesPerPixel        = 1
    ds.PhotometricInterpretation = "MONOCHROME2"
    
    bits = arr.itemsize * 8
    ds.BitsAllocated       = ds.BitsStored = bits
    ds.HighBit             = bits - 1
    ds.PixelRepresentation = int(arr.dtype.kind == "i")    
    
    insert_all_creators(ds)                       # Private Creator SQ

    # 3) SQ 태그 선(先) 처리 ------------------------------------------
    for _, sq in tags[tags["vr"] == "SQ"].iterrows():
        seq = build_sequence(sq["group"] + sq["element"], tags)
        if seq:
            ds.add_new((int(sq["group"],16),int(sq["element"],16)),"SQ",seq)

 # ─────────────────── 이미지·시스템 기본값 보완 ───────────────────
    tag_map = {(r["group"], r["element"]): r for _, r in tags.iterrows()}
    img_defaults = { ("0028","0010"): rows, ("0028","0011"): cols,
        ("0028","0100"): bits, ("0028","0101"): bits, ("0028","0102"): bits-1,
        ("0028","0103"): ds.PixelRepresentation, ("0028","0002"): 1,
        ("0028","0004"): "MONOCHROME2" }
    now = dt.now()
    sys_defaults = { ("0008","0060"): "OT",
        ("0008","0020"): now.strftime("%Y%m%d"),
        ("0008","0030"): now.strftime("%H%M%S"),
        ("0020","000D"): ds.StudyInstanceUID,
        ("0020","000E"): ds.SeriesInstanceUID,
        ("0020","0052"): generate_uid(prefix=OID_PREFIX) }
    for (g,e), val in {**img_defaults, **sys_defaults}.items():
        if (g,e) not in tag_map or tag_map[(g,e)]["value"] == "":
            vr = VR_DEF[(g,e)]
            ds.add_new((int(g,16),int(e,16)), vr, safe_value(vr,val))


    # ── 4) 일반 태그 처리 ────────────────────────────────────────────
    # ── 4) 일반 태그 처리 ────────────────────────────────────────────
    parent_ds = ds                              # 기본 대상 = 루트
    for _, r in tags.iterrows():
        pt = r.get("parenttag", "")             # parenttag 없으면 공란

        if r["vr"] == "SQ":
            continue
        if not r["group"] or not r["element"]:
            LOG.warning("skip blank G/E row : %s", r.to_dict())
            continue

        # ── parenttag 판별 ───────────────────────────────────────
        if pt:                                  # parenttag 가 있는 경우
            if pt.endswith("0000"):             # Group Length → 루트
                parent_ds = ds
            else:
                g_p = _hex2int(pt[:4]); e_p = _hex2int(pt[4:])
                if g_p < 0 or e_p < 0:
                    continue
                if (g_p, e_p) not in ds:        # 상위 SQ 없으면 생성
                    ds.add_new((g_p, e_p), "SQ", [Dataset()])
                seq_elem = ds[(g_p, e_p)]

                # SQ 여부 검사
                if seq_elem.VR != "SQ":
                    LOG.warning("%04X,%04X VR=%s → root",
                                g_p, e_p, seq_elem.VR)
                    parent_ds = ds
                    continue                    # ← continue 는 여기서만!

                # Item 재사용
                if (not seq_elem.value or
                        not isinstance(seq_elem.value[-1], Dataset)):
                    seq_elem.value.append(Dataset())
                parent_ds = seq_elem.value[-1]
        else:
            parent_ds = ds                      # parenttag 공란 → 루트

        # ── 태그 삽입 ───────────────────────────────────────────
        g = _hex2int(r["group"]); e = _hex2int(r["element"])
        if g < 0 or e < 0:
            continue

        # 0002 그룹은 File-Meta 로
        target = meta if g == 0x0002 else parent_ds
        if (g == 0x0002) and (g, e) in target:
            continue                            # 중복 방지

        target.add_new(
            (g, e),
            VR_DOWN.get(r["vr"], r["vr"]),
            safe_value(r["vr"], r["value"])
        )



    # 5) Rescale & PixelData -----------------------------------------
    slope, intercept = get_rescale_params(tags, gui_slope, gui_int, override)
    ds.RescaleSlope, ds.RescaleIntercept, ds.RescaleType = slope, intercept, "HU"
    ds.PixelData = arr.tobytes()

    return ds



# ── 간단 태그 검사 ───────────────────────────────────
def validate_tags(df) -> List[Dict[str,str]]:
    issues = []
    # DA 오류 처리
    bad_da = df[(df["vr"]=="DA") & (~df["value"].str.fullmatch(r"\d{8}", na=False))]
    for _, r in bad_da.iterrows():
        issues.append({
            "group": r["group"],
            "element": r["element"],
            "vr": r["vr"],
            "value": r["value"],
            "error": f"DA 오류: 값 {r['value']}은(는) YYYYMMDD 형식이 아닙니다."
        })
    # TM 오류 처리
    bad_tm = df[(df["vr"]=="TM") & (~df["value"].str.fullmatch(r"\d{6}", na=False))]
    for _, r in bad_tm.iterrows():
        issues.append({
            "group": r["group"],
            "element": r["element"],
            "vr": r["vr"],
            "value": r["value"],
            "error": f"TM 오류: 값 {r['value']}은(는) HHMMSS 형식이 아닙니다."
        })
    return issues
# ── manifest URL 탐색 ───────────────────────────────
def find_manifest_url(df: pd.DataFrame) -> str:
    cand = df[(df["keyword"].str.contains("IIIF", case=False, na=False)) &
              (df["value"].str.contains("http", na=False))]
    return cand["value"].iloc[0] if not cand.empty else ""

# ── GUI ──────────────────────────────────────────────
class ConverterGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("DICOCH DICOM Converter v3.2")
        self.geometry("1020x780")
        self._build()

    # ── GUI 레이아웃 ────────────────────────────────
    def _build(self):
            # Notebook ────────────────────────
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)

        converter_tab = ttk.Frame(notebook)
        info_tab      = ttk.Frame(notebook)
        notebook.add(converter_tab, text="Converter")
        notebook.add(info_tab,     text="Info")

        # ── 1) Converter 탭 ───────────────────────────
        frm = ttk.Frame(converter_tab, padding=12)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(1, weight=1)

        # ── 입력/출력 경로
        self.e_in, self.e_tag, self.e_out = [ttk.Entry(frm) for _ in range(3)]
        ttk.Label(frm, text="TIFF Folder (Image Files):").grid(row=0, column=0, sticky="w")
        self.e_in.grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Button(frm, text="Browse", command=self._pick_in)\
            .grid(row=0, column=2)

        ttk.Label(frm, text="Tag Information (Excel File):").grid(row=1, column=0, sticky="w")
        self.e_tag.grid(row=1, column=1, sticky="ew", padx=4)
        ttk.Button(frm, text="Browse", command=self._pick_tag)\
            .grid(row=1, column=2)

        ttk.Label(frm, text="Output Folder:").grid(row=2, column=0, sticky="w")
        self.e_out.grid(row=2, column=1, sticky="ew", padx=4)
        ttk.Button(frm, text="Browse", command=self._pick_out)\
            .grid(row=2, column=2)

        # Slope / Intercept
        ttk.Label(frm, text="Slope:").grid(row=0, column=3, sticky="e")
        self.e_slope = ttk.Entry(frm, width=8); self.e_slope.insert(0, "1")
        self.e_slope.grid(row=0, column=4)
        ttk.Label(frm, text="Intercept:").grid(row=1, column=3, sticky="e")
        self.e_int   = ttk.Entry(frm, width=8); self.e_int.insert(0, "-1024")
        self.e_int.grid(row=1, column=4)


        self.gui_override = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm, text="Apply Slope/Intercept in GUI",
                        variable=self.gui_override)\
        .grid(row=2, column=3, columnspan=2, sticky="w", pady=2)

        # IIIF 옵션
        self.open_viewer = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm, text="Open IIIF Viewer", variable=self.open_viewer)\
            .grid(row=3, column=3, sticky="w")
        ttk.Label(frm, text="Manifest URL (Optional):").grid(row=3, column=0, sticky="w")
        self.e_manifest = ttk.Entry(frm)
        self.e_manifest.grid(row=3, column=1, columnspan=2, sticky="ew", padx=4)

        # 자동 폴더 열기
        self.auto_open = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm, text="Open Output Folder", variable=self.auto_open)\
            .grid(row=4, column=3, columnspan=2, sticky="w")

        # 버튼 영역
        bf = ttk.Frame(frm); bf.grid(row=4, column=0, columnspan=3, sticky="ew", pady=10)
        ttk.Button(bf, text="Convert to DCM", command=self._start).pack(side="left", expand=True, fill="x", padx=(0,4))
        ttk.Button(bf, text="Validate Tags", command=self._check).pack(side="left", expand=True, fill="x")
        ttk.Button(bf, text="Save Tag Information", command=self._save_tags).pack(side="left", expand=True, fill="x")

        # 진행률 / 로그
        self.pb  = ttk.Progressbar(frm); self.pb.grid(row=5, column=0, columnspan=5, sticky="ew")
        ttk.Label(frm, text="■ Data Processing Status").grid(row=6, column=0, sticky="w")
        self.log = scrolledtext.ScrolledText(frm, height=10)
        self.log.grid(row=7, column=0, columnspan=5, sticky="nsew", pady=4)

        ttk.Label(frm, text="■ Tag Processing Result").grid(row=8, column=0, sticky="w")
        self.tag_view = scrolledtext.ScrolledText(frm, height=12,
                                                font=("Consolas", 9),
                                                padx=2, pady=0, wrap="none")
        self.tag_view.grid(row=9, column=0, columnspan=5, sticky="nsew")
        frm.rowconfigure(9, weight=1)

        # ── 2) Info 탭 ────────────────────────────────
        INFO_TEXT = (
        "DICOCH DICOM Converter  v2.9s  (2025-06-20)\n\n"

        "▶ 주요 업데이트 │ Updates\n"
        "1) 16-byte Creator token + FullName(UT)\n"
        "2) Tag viewer & save  (TXT / XLSX)\n"
        "3) IIIF auto-viewer option\n\n"

        "▶ 사용법 │ How to use\n"
        "1) Converter 탭에서 TIFF·엑셀·출력 폴더 지정\n"
        "2) [변환 시작] → DICOM 생성\n"
        "3) 필요 시 [IIIF 뷰어 열기] 체크\n"
        "4) 태그 결과 저장 버튼으로 TXT/XLSX 추출\n\n"

        "▶ 제작자 │ Author\n"
        "▶ 제작자 │ Author\n"
        "기관  : 국립문화재연구원 X-선·CT 분석실\n"
        "        National Research Institute of Cultural Heritage (NRICH), X-ray / CT Lab\n"
        "업무  : 문화유산 X-선·CT 비파괴 진단 · 3D 스캔 · 연륜연대 분석 · 디지털 데이터 표준 연구\n"
        "        Non-destructive X-ray/CT diagnostics, 3D scanning & dendrochronology, digital-standard research\n"
        "이름  : 송정일  Song Jung-il\n"
        "e-mail: ssong85@korea.kr\n"
        "배포  : CC-BY-SA 4.0 — 자유 복제·수정·재배포(출처 표기)\n"
        "        CC-BY-SA 4.0 — Free use, modification and redistribution (attribution required)\n"
        "GitHub: https://github.com/SONG-JUNG-IL/DICOCH_TAG\n"
        )
        info_box = scrolledtext.ScrolledText(
            info_tab,
            font=("Segoe UI", 11),
            wrap="word",
            padx=12, pady=10,
            state="normal"
        )
        info_box.insert("1.0", INFO_TEXT)

        # 헤더(▶ …) 줄만 굵게
        for line in (3, 9, 17):          # 1-based line 번호
            info_box.tag_add("hdr", f"{line}.0", f"{line}.end")
        info_box.tag_configure("hdr", font=("Segoe UI", 11, "bold"))

        # 전체 행간 살짝 여유
        info_box.tag_add("gap", "1.0", "end")
        info_box.tag_configure("gap", spacing1=2, spacing3=2)

        info_box.config(state="disabled")
        info_box.pack(fill="both", expand=True)
    # ── 파일선택 유틸 ────────────────────────────────
    def _pick_in(self):
        p=filedialog.askdirectory()
        if p:
            self.e_in.delete(0,tk.END); self.e_in.insert(0,p)
            ts=dt.now().strftime("%Y%m%d_%H%M%S")
            self.e_out.delete(0,tk.END); self.e_out.insert(0,str(Path.cwd()/f"{Path(p).name}_{ts}"))

    def _pick_tag(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if p:
            self.e_tag.delete(0,tk.END); self.e_tag.insert(0,p)

    def _pick_out(self):
        base=filedialog.askdirectory()
        if base:
            ts=dt.now().strftime("%Y%m%d_%H%M%S")
            self.e_out.delete(0,tk.END); self.e_out.insert(0,str(Path(base)/f"output_{ts}"))

    # ── 로그 출력 도우미
    def _log(self,m): self.log.insert(tk.END,m+"\n"); self.log.see(tk.END)

    # ── 태그 검사
    def _check(self):
        try:
            tags    = load_tags(Path(self.e_tag.get()))
            issues  = validate_tags(tags)
            if issues:
                for it in issues:
                    msg = f"✖ {it['group']}:{it['element']} → {it['error']}"
                    self._log(msg)
                # GUI 저장용: 태그 DataFrame + 오류 리스트 보관
                self.cur_tag_errors = pd.DataFrame(issues)                
                messagebox.showwarning("Validation",f"Issue Detected {len(issues)} Count")
            else:
                self._log("No Tag Structure Issues")
                messagebox.showinfo("Validation","Normal")
        except Exception as e:
            self._log(str(e)); messagebox.showerror("Error",str(e))

    # ── 변환 스레드 시작
    def _start(self):
        self.pb.config(value=0)
        threading.Thread(target=self._convert,daemon=True).start()

    def _update_tag_view(self, df: pd.DataFrame) -> None:
        """
        태그 DataFrame(df)을 고정폭 문자열로 변환해 self.tag_view 위젯에 표시.
        • 행간 0  • 수평 스크롤 지원
        """
        # 1) 컬럼별 최소 폭 = max(헤더, 값 최대) + 2
        widths = {c: max(len(c), df[c].astype(str).map(len).max()) + 2
                for c in df.columns}
        pretty = df.to_string(index=False, col_space=widths, justify="left")

        # 2) 기존 위젯 내용만 갱신
        self.tag_view.config(state="normal")
        self.tag_view.delete("1.0", tk.END)
        self.tag_view.insert(tk.END, pretty)

        # 3) 행간·스크롤 설정
        self.tag_view.tag_add("tight", "1.0", "end")
        self.tag_view.tag_configure("tight", spacing1=0, spacing3=0)
        self.tag_view.config(state="disabled")



    # ── 태그 결과 저장
    # ── 태그 결과 저장 ────────────────────────────────────────────
    def _save_tags(self):
        if not hasattr(self, "cur_tags"):
            messagebox.showwarning("저장", "변환된 태그 정보가 없습니다.")
            return

        out_dir = Path(self.e_out.get())
        if not out_dir.exists():
            messagebox.showerror("저장", "출력 폴더가 없습니다.")
            return

        ts = dt.now().strftime("%Y%m%d_%H%M%S")
        txt_path  = out_dir / f"tag_info_{ts}.txt"
        xlsx_path = out_dir / f"tag_info_{ts}.xlsx"
        json_path = out_dir / f"tag_info_{ts}.json"

        # ── ① 에러 컬럼 병합
        tags_df = self.cur_tags.copy()
        if hasattr(self, "cur_tag_errors") and not self.cur_tag_errors.empty:
            # group+element 키 생성
            errs = self.cur_tag_errors.copy()
            errs["key"] = errs["group"] + errs["element"]
            tags_df["key"] = tags_df["group"] + tags_df["element"]
            # left-join, 없는 항목은 NaN → 빈 문자열
            tags_df = (tags_df
                       .merge(errs[["key", "error"]], on="key", how="left")
                       .drop(columns=["key"]))
            tags_df["error"] = tags_df["error"].fillna("")
        else:
            tags_df["error"] = ""

        # ── ② 파일로 저장
        tags_df.to_csv(txt_path, sep="\t", index=False)
        tags_df.to_excel(xlsx_path, index=False)
        data = tags_df.to_dict(orient="records")
        json_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

        # ── ③ 로그 및 알림
        self._log(f"[Tag Save] {txt_path.name}, {xlsx_path.name}, {json_path.name}")
        messagebox.showinfo(
            "저장 완료",
            f"TXT / XLSX / JSON 저장\n{txt_path}\n{xlsx_path}\n{json_path}"
        )


    # ── 변환 로직
    def _convert(self):
        try:
            in_dir = Path(self.e_in.get()); tag_xls = Path(self.e_tag.get()); out_dir = Path(self.e_out.get())
            if not in_dir.is_dir() or not tag_xls.is_file():
                messagebox.showerror("오류","경로 확인"); return
            tiffs=sorted(in_dir.glob("*.tif*"))
            if not tiffs:
                messagebox.showerror("오류","TIFF 없음"); return

            gui_slope=float(self.e_slope.get() or 1)
            gui_int =float(self.e_int.get() or -1024)

            tags=load_tags(tag_xls)         # 태그 로드
            self.cur_tags = tags            # GUI 저장용
            self.after(0, lambda: self._update_tag_view(tags))

            # ➊ 출력 폴더
            out_dir.mkdir(parents=True, exist_ok=True)
            # ➋ 원본 .dic 작성
            tmp_path = write_dic(tags, out_dir)
            # ➌ 원하는 .txt 이름으로 변경  ★★ 새로 삽입
            dic_fname = f"Dic_DICOCH_DICOM_PravateTag_{dt.now():%Y%m%d_%H%M%S}.txt"   
            dic_path  = tmp_path.with_name(dic_fname)
            tmp_path.rename(dic_path)
            # ➍ 로그 기록 라인 수정  ★★
            self._log(f"[{dic_fname}] {dic_path}")
            
            manifest_val = find_manifest_url(tags)
            if manifest_val:
                self._log(f"[IIIF Link] {manifest_val}")

            log_f=(out_dir/f"log_{dt.now():%Y%m%d_%H%M%S}.txt").open("w",encoding="utf-8")
            if manifest_val: log_f.write(f"[IIIF Link] {manifest_val}\n")

            succ=fail=0; lock=threading.Lock(); self.pb.config(maximum=len(tiffs))

            def task(fp:Path):
                nonlocal succ,fail
                try:
                    ds = build_dataset(fp, tags, gui_slope, gui_int, override=self.gui_override.get())
                    pydicom.dcmwrite(out_dir / f"{fp.stem}.dcm", ds, write_like_original=False)
                    with lock: succ+=1
                    return f"✔ {fp.name}"
                except Exception as e:
                    with lock: fail+=1
                    return f"✖ {fp.name} → {e}"

            with cf.ThreadPoolExecutor(max_workers=max(1, os.cpu_count()//2)) as ex:
                for idx,msg in enumerate(ex.map(task,tiffs),1):
                    self._log(msg); log_f.write(msg+"\n"); self.pb.config(value=idx)

            summary=f"Completed {succ}   Failed {fail}"
            self._log(summary); log_f.write(summary+"\n"); log_f.close()
            messagebox.showinfo("Completed",summary)

            # 폴더 자동 열기
            if self.auto_open.get() and succ and not fail:
                try: os.startfile(out_dir)
                except AttributeError:
                    opener="open" if sys.platform=="darwin" else "xdg-open"
                    subprocess.Popen([opener,str(out_dir)])

            # IIIF 뷰어 호출
            if self.open_viewer.get():
                manifest = self.e_manifest.get().strip() or manifest_val
                if manifest:
                    url = MIRADOR_DEMO + manifest   # no encoding
                    self._log(f"[Mirador] {url}")
                    webbrowser.open_new_tab(url)
                else:
                    self._log("Manifest URL이 없어 IIIF 뷰어 호출 생략")

        finally:
            self.pb.config(value=0)

# ── 실행 ──────────────────────────────────────────────
def main():
    try:
        import ctypes; ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    ConverterGUI().mainloop()

if __name__=="__main__":
    main()
